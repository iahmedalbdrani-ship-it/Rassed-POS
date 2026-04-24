"""
=============================================================================
  Raseed (رصيد) — Multi-Role Accounting Data Exporter
  Exports: Clients, Cashiers, Supervisors, Owner Summary
  Output:  reports/ directory with structured Excel files
=============================================================================
"""

# ── Standard library ────────────────────────────────────────────────────────
import os
import sys
from datetime import datetime
from pathlib import Path

# ── Third-party ──────────────────────────────────────────────────────────────
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from sqlalchemy import create_engine, text
except ImportError as e:
    sys.exit(
        f"\n[ERROR] Missing dependency: {e}\n"
        "Install with:\n"
        "  pip install pandas openpyxl sqlalchemy\n"
        "  # For PostgreSQL: pip install psycopg2-binary\n"
        "  # For MySQL:      pip install pymysql\n"
        "  # For SQL Server: pip install pyodbc\n"
    )

# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                     DATABASE CONFIGURATION                              ║
# ║  Fill in your connection details here before running                    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

DB_CONFIG = {
    # ── Choose ONE of the connection strings below and fill it in ──────────
    # PostgreSQL
    "url": "postgresql+psycopg2://USER:PASSWORD@HOST:5432/DATABASE",

    # MySQL / MariaDB
    # "url": "mysql+pymysql://USER:PASSWORD@HOST:3306/DATABASE",

    # SQL Server
    # "url": "mssql+pyodbc://USER:PASSWORD@HOST/DATABASE?driver=ODBC+Driver+17+for+SQL+Server",

    # SQLite (for testing with a local file)
    # "url": "sqlite:///raseed.db",
}

# ── Table names (adjust if your schema differs) ────────────────────────────
TABLES = {
    "clients":      "clients",
    "cashiers":     "cashiers",
    "supervisors":  "supervisors",
    "transactions": "transactions",
    "accounts":     "accounts",
    "owners":       "owners",
}

# ── Output root ────────────────────────────────────────────────────────────
OUTPUT_ROOT = Path("reports")


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                         DESIGN CONSTANTS                                ║
# ╚══════════════════════════════════════════════════════════════════════════╝

# Color palette (hex strings for openpyxl)
CLR = {
    "header_bg":    "1E3A5F",   # Dark navy – header row background
    "header_fg":    "FFFFFF",   # White text on headers
    "subheader_bg": "2E86AB",   # Blue – sub-headers / group rows
    "accent1":      "A8DADC",   # Light teal – alternating rows
    "accent2":      "F1FAEE",   # Near-white – alternating rows
    "positive":     "1B4332",   # Dark green text – credit / income
    "negative":     "7B2D2D",   # Dark red text – debit / expense
    "positive_bg":  "D8F3DC",   # Green fill – positive cells
    "negative_bg":  "FFE0E0",   # Red fill – negative cells
    "total_bg":     "343A40",   # Dark grey – totals row background
    "total_fg":     "FFFFFF",   # White text on totals row
    "title_fg":     "1E3A5F",   # Navy – sheet title text
    "border":       "ADB5BD",   # Light grey borders
}

FONT_NAME = "Arial"


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                         STYLING HELPERS                                 ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _border(color: str = CLR["border"]) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)


def style_header_row(ws, row: int, num_cols: int):
    """Apply dark-navy header styling to a specific row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _font(bold=True, color=CLR["header_fg"], size=11)
        cell.fill = _fill(CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()


def style_data_row(ws, row: int, num_cols: int, alternate: bool = False):
    """Zebra-stripe data rows."""
    bg = CLR["accent1"] if alternate else CLR["accent2"]
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(bg)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = _font(size=10)


def style_totals_row(ws, row: int, num_cols: int):
    """Dark total / summary row at the bottom."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _font(bold=True, color=CLR["total_fg"], size=11)
        cell.fill = _fill(CLR["total_bg"])
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _border()


def add_sheet_title(ws, title: str, subtitle: str = "", cols: int = 6):
    """Merge cells and write a styled title block at the top."""
    ws.insert_rows(1, amount=3)

    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name=FONT_NAME, bold=True, size=16, color=CLR["title_fg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = _fill("EAF4FB")
    ws.row_dimensions[1].height = 32

    if subtitle:
        ws.merge_cells(f"A2:{get_column_letter(cols)}2")
        sub_cell = ws["A2"]
        sub_cell.value = subtitle
        sub_cell.font = Font(name=FONT_NAME, size=10, color="666666", italic=True)
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        sub_cell.fill = _fill("F8FAFB")
        ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6  # spacer


def set_column_widths(ws, widths: dict):
    """widths = {"A": 20, "B": 14, ...}"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def format_currency_col(ws, col: int, start_row: int, end_row: int, red_if_negative: bool = True):
    """Apply SAR number format to a column range."""
    fmt = '#,##0.00 "ر.س"'
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        cell.number_format = fmt
        if red_if_negative and isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.font = _font(color=CLR["negative"])
            cell.fill = _fill(CLR["negative_bg"])


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                       DATABASE UTILITIES                                ║
# ╚══════════════════════════════════════════════════════════════════════════╝

_engine = None


def get_engine():
    global _engine
    if _engine is None:
        _engine = create_engine(DB_CONFIG["url"], pool_pre_ping=True)
    return _engine


def query_df(sql: str, params: dict = None) -> pd.DataFrame:
    """Run a SELECT query and return a DataFrame (empty DF on error)."""
    try:
        with get_engine().connect() as conn:
            return pd.read_sql(text(sql), conn, params=params or {})
    except Exception as exc:
        print(f"  [WARN] Query failed: {exc}")
        return pd.DataFrame()


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                        FOLDER STRUCTURE                                 ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def create_output_dirs():
    for sub in ("clients", "cashiers", "supervisors", "owner"):
        (OUTPUT_ROOT / sub).mkdir(parents=True, exist_ok=True)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                      1. CLIENTS EXPORTER                                ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def export_clients() -> int:
    """Export one Excel file per client with a transaction ledger."""
    print("\n── Exporting Clients ─────────────────────────────────────────")

    clients_df = query_df(f"SELECT * FROM {TABLES['clients']} ORDER BY id")
    if clients_df.empty:
        print("  No clients found.")
        return 0

    # Detect id / name columns flexibly
    id_col   = next((c for c in clients_df.columns if "id"   in c.lower()), clients_df.columns[0])
    name_col = next((c for c in clients_df.columns if "name" in c.lower()), clients_df.columns[1] if len(clients_df.columns) > 1 else id_col)

    count = 0
    for _, client in clients_df.iterrows():
        client_id   = client[id_col]
        client_name = str(client[name_col]).replace(" ", "_")

        # ── Fetch transactions ────────────────────────────────────────────
        txn_df = query_df(
            f"""
            SELECT
                transaction_date  AS "Date",
                description       AS "Description",
                debit_amount      AS "Debit",
                credit_amount     AS "Credit"
            FROM {TABLES['transactions']}
            WHERE client_id = :cid
            ORDER BY transaction_date ASC
            """,
            {"cid": client_id},
        )

        # ── Build workbook ────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Account Statement"

        num_cols = 5
        add_sheet_title(
            ws,
            f"Account Statement — {client[name_col]}",
            f"Client ID: {client_id}  |  Generated: {datetime.now():%Y-%m-%d %H:%M}",
            cols=num_cols,
        )

        # Headers (row 4 after title block + spacer)
        headers = ["Date", "Description", "Debit (ر.س)", "Credit (ر.س)", "Running Balance (ر.س)"]
        header_row = 4
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=h)
        style_header_row(ws, header_row, num_cols)

        # Data rows
        running_balance = 0.0
        data_start = header_row + 1

        if txn_df.empty:
            ws.cell(row=data_start, column=1, value="No transactions found for this client.")
            ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
        else:
            for r_idx, (_, txn) in enumerate(txn_df.iterrows(), start=data_start):
                debit  = float(txn.get("Debit")  or 0)
                credit = float(txn.get("Credit") or 0)
                running_balance += credit - debit

                row_vals = [
                    txn.get("Date"),
                    txn.get("Description", ""),
                    debit,
                    credit,
                    running_balance,
                ]
                for c_idx, val in enumerate(row_vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

                style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))

                # Colour debit/credit cells
                if debit > 0:
                    ws.cell(row=r_idx, column=3).fill = _fill(CLR["negative_bg"])
                    ws.cell(row=r_idx, column=3).font = _font(color=CLR["negative"])
                if credit > 0:
                    ws.cell(row=r_idx, column=4).fill = _fill(CLR["positive_bg"])
                    ws.cell(row=r_idx, column=4).font = _font(color=CLR["positive"])

                # Running balance colour
                bal_cell = ws.cell(row=r_idx, column=5)
                if running_balance < 0:
                    bal_cell.fill = _fill(CLR["negative_bg"])
                    bal_cell.font = _font(bold=True, color=CLR["negative"])
                else:
                    bal_cell.fill = _fill(CLR["positive_bg"])
                    bal_cell.font = _font(bold=True, color=CLR["positive"])

                # Date & number formats
                ws.cell(row=r_idx, column=1).number_format = "YYYY-MM-DD"
                for fc in [3, 4, 5]:
                    ws.cell(row=r_idx, column=fc).number_format = '#,##0.00 "ر.س"'

            # Totals row
            last_data = data_start + len(txn_df) - 1
            totals_row = last_data + 1
            total_debit  = txn_df["Debit"].fillna(0).astype(float).sum()
            total_credit = txn_df["Credit"].fillna(0).astype(float).sum()

            ws.cell(row=totals_row, column=1, value="TOTAL")
            ws.cell(row=totals_row, column=2, value="")
            ws.cell(row=totals_row, column=3, value=f'=SUM(C{data_start}:C{last_data})')
            ws.cell(row=totals_row, column=4, value=f'=SUM(D{data_start}:D{last_data})')
            ws.cell(row=totals_row, column=5, value=running_balance)
            style_totals_row(ws, totals_row, num_cols)
            for fc in [3, 4, 5]:
                ws.cell(row=totals_row, column=fc).number_format = '#,##0.00 "ر.س"'

        set_column_widths(ws, {"A": 14, "B": 40, "C": 18, "D": 18, "E": 22})

        # ── Save ──────────────────────────────────────────────────────────
        safe_name = "".join(c if c.isalnum() or c in "_-" else "_" for c in client_name)
        out_path = OUTPUT_ROOT / "clients" / f"{client_id}_{safe_name}.xlsx"
        wb.save(out_path)
        print(f"  ✓ {out_path.name}")
        count += 1

    return count


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                      2. CASHIERS EXPORTER                               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def export_cashiers() -> int:
    print("\n── Exporting Cashiers ────────────────────────────────────────")

    cashiers_df = query_df(f"SELECT * FROM {TABLES['cashiers']} ORDER BY id")
    if cashiers_df.empty:
        print("  No cashiers found.")
        return 0

    id_col   = next((c for c in cashiers_df.columns if "id"   in c.lower()), cashiers_df.columns[0])
    name_col = next((c for c in cashiers_df.columns if "name" in c.lower()), cashiers_df.columns[1] if len(cashiers_df.columns) > 1 else id_col)

    count = 0
    for _, cashier in cashiers_df.iterrows():
        cashier_id   = cashier[id_col]
        cashier_name = str(cashier[name_col]).replace(" ", "_")

        txn_df = query_df(
            f"""
            SELECT
                transaction_date                             AS "Date",
                shift                                        AS "Shift",
                SUM(CASE WHEN txn_type='collection' THEN amount ELSE 0 END) AS "Total Collections",
                SUM(CASE WHEN txn_type='payment'    THEN amount ELSE 0 END) AS "Total Payments",
                SUM(CASE WHEN txn_type='collection' THEN amount ELSE -amount END) AS "Net Cash"
            FROM {TABLES['transactions']}
            WHERE cashier_id = :cid
            GROUP BY transaction_date, shift
            ORDER BY transaction_date ASC, shift ASC
            """,
            {"cid": cashier_id},
        )

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Daily Transactions"
        num_cols = 5

        add_sheet_title(
            ws,
            f"Cashier Report — {cashier[name_col]}",
            f"Cashier ID: {cashier_id}  |  Generated: {datetime.now():%Y-%m-%d %H:%M}",
            cols=num_cols,
        )

        headers    = ["Date", "Shift", "Total Collections (ر.س)", "Total Payments (ر.س)", "Net Cash (ر.س)"]
        header_row = 4
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=h)
        style_header_row(ws, header_row, num_cols)

        data_start = header_row + 1

        if txn_df.empty:
            ws.cell(row=data_start, column=1, value="No transactions found for this cashier.")
            ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
        else:
            for r_idx, (_, row) in enumerate(txn_df.iterrows(), start=data_start):
                net = float(row.get("Net Cash") or 0)
                vals = [
                    row.get("Date"),
                    row.get("Shift", ""),
                    float(row.get("Total Collections") or 0),
                    float(row.get("Total Payments") or 0),
                    net,
                ]
                for c_idx, val in enumerate(vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
                style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))

                ws.cell(row=r_idx, column=1).number_format = "YYYY-MM-DD"
                for fc in [3, 4, 5]:
                    ws.cell(row=r_idx, column=fc).number_format = '#,##0.00 "ر.س"'

                net_cell = ws.cell(row=r_idx, column=5)
                if net < 0:
                    net_cell.fill = _fill(CLR["negative_bg"])
                    net_cell.font = _font(bold=True, color=CLR["negative"])
                else:
                    net_cell.fill = _fill(CLR["positive_bg"])
                    net_cell.font = _font(bold=True, color=CLR["positive"])

            last_data  = data_start + len(txn_df) - 1
            totals_row = last_data + 1
            ws.cell(row=totals_row, column=1, value="TOTAL")
            ws.cell(row=totals_row, column=2, value="")
            ws.cell(row=totals_row, column=3, value=f'=SUM(C{data_start}:C{last_data})')
            ws.cell(row=totals_row, column=4, value=f'=SUM(D{data_start}:D{last_data})')
            ws.cell(row=totals_row, column=5, value=f'=SUM(E{data_start}:E{last_data})')
            style_totals_row(ws, totals_row, num_cols)
            for fc in [3, 4, 5]:
                ws.cell(row=totals_row, column=fc).number_format = '#,##0.00 "ر.س"'

        set_column_widths(ws, {"A": 14, "B": 16, "C": 24, "D": 22, "E": 20})

        safe_name = "".join(c if c.isalnum() or c in "_-" else "_" for c in cashier_name)
        out_path  = OUTPUT_ROOT / "cashiers" / f"{cashier_id}_{safe_name}.xlsx"
        wb.save(out_path)
        print(f"  ✓ {out_path.name}")
        count += 1

    return count


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                    3. SUPERVISORS EXPORTER                              ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def export_supervisors() -> int:
    print("\n── Exporting Supervisors ─────────────────────────────────────")

    sup_df = query_df(f"SELECT * FROM {TABLES['supervisors']} ORDER BY id")
    if sup_df.empty:
        print("  No supervisors found.")
        return 0

    id_col   = next((c for c in sup_df.columns if "id"   in c.lower()), sup_df.columns[0])
    name_col = next((c for c in sup_df.columns if "name" in c.lower()), sup_df.columns[1] if len(sup_df.columns) > 1 else id_col)

    count = 0
    for _, sup in sup_df.iterrows():
        sup_id   = sup[id_col]
        sup_name = str(sup[name_col]).replace(" ", "_")

        activity_df = query_df(
            f"""
            SELECT
                action_date        AS "Date",
                action_type        AS "Action",
                approved_by        AS "Approved By",
                amount             AS "Amount",
                status             AS "Status"
            FROM {TABLES['transactions']}
            WHERE supervisor_id = :sid
            ORDER BY action_date ASC
            """,
            {"sid": sup_id},
        )

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Activity Log"
        num_cols = 5

        add_sheet_title(
            ws,
            f"Supervisor Activity — {sup[name_col]}",
            f"Supervisor ID: {sup_id}  |  Generated: {datetime.now():%Y-%m-%d %H:%M}",
            cols=num_cols,
        )

        headers    = ["Date", "Action", "Approved By", "Amount (ر.س)", "Status"]
        header_row = 4
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=h)
        style_header_row(ws, header_row, num_cols)

        data_start = header_row + 1

        if activity_df.empty:
            ws.cell(row=data_start, column=1, value="No activity found for this supervisor.")
            ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
        else:
            for r_idx, (_, row) in enumerate(activity_df.iterrows(), start=data_start):
                status = str(row.get("Status", "")).lower()
                vals = [
                    row.get("Date"),
                    row.get("Action", ""),
                    row.get("Approved By", ""),
                    float(row.get("Amount") or 0),
                    row.get("Status", ""),
                ]
                for c_idx, val in enumerate(vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
                style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))

                ws.cell(row=r_idx, column=1).number_format = "YYYY-MM-DD"
                ws.cell(row=r_idx, column=4).number_format = '#,##0.00 "ر.س"'

                # Status badge colour
                status_cell = ws.cell(row=r_idx, column=5)
                if "approved" in status:
                    status_cell.fill = _fill(CLR["positive_bg"])
                    status_cell.font = _font(bold=True, color=CLR["positive"])
                elif "rejected" in status or "denied" in status:
                    status_cell.fill = _fill(CLR["negative_bg"])
                    status_cell.font = _font(bold=True, color=CLR["negative"])
                elif "pending" in status:
                    status_cell.fill = _fill("FFF3CD")
                    status_cell.font = _font(bold=True, color="856404")

            last_data  = data_start + len(activity_df) - 1
            totals_row = last_data + 1
            ws.cell(row=totals_row, column=1, value="TOTAL")
            ws.cell(row=totals_row, column=2, value=f"{len(activity_df)} Actions")
            ws.cell(row=totals_row, column=3, value="")
            ws.cell(row=totals_row, column=4, value=f'=SUM(D{data_start}:D{last_data})')
            ws.cell(row=totals_row, column=5, value="")
            style_totals_row(ws, totals_row, num_cols)
            ws.cell(row=totals_row, column=4).number_format = '#,##0.00 "ر.س"'

        set_column_widths(ws, {"A": 14, "B": 30, "C": 22, "D": 20, "E": 16})

        safe_name = "".join(c if c.isalnum() or c in "_-" else "_" for c in sup_name)
        out_path  = OUTPUT_ROOT / "supervisors" / f"{sup_id}_{safe_name}.xlsx"
        wb.save(out_path)
        print(f"  ✓ {out_path.name}")
        count += 1

    return count


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                   4. OWNER MASTER REPORT                                ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _sheet_balance_summary(wb: Workbook):
    """Sheet 1 — Overall Balance Summary."""
    ws = wb.create_sheet("Overall Balance")

    summary_df = query_df(
        f"""
        SELECT
            account_type        AS "Account Type",
            SUM(debit_amount)   AS "Total Debits",
            SUM(credit_amount)  AS "Total Credits",
            SUM(credit_amount - debit_amount) AS "Net Balance"
        FROM {TABLES['transactions']}
        JOIN {TABLES['accounts']} ON {TABLES['transactions']}.account_id = {TABLES['accounts']}.id
        GROUP BY account_type
        ORDER BY account_type
        """
    )

    num_cols = 4
    add_sheet_title(ws, "Overall Balance Summary", f"Generated: {datetime.now():%Y-%m-%d %H:%M}", cols=num_cols)

    headers    = ["Account Type", "Total Debits (ر.س)", "Total Credits (ر.س)", "Net Balance (ر.س)"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, header_row, num_cols)

    data_start = header_row + 1

    if summary_df.empty:
        ws.cell(row=data_start, column=1, value="No account data found.")
        ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
    else:
        for r_idx, (_, row) in enumerate(summary_df.iterrows(), start=data_start):
            net = float(row.get("Net Balance") or 0)
            vals = [
                row.get("Account Type", ""),
                float(row.get("Total Debits")  or 0),
                float(row.get("Total Credits") or 0),
                net,
            ]
            for c_idx, val in enumerate(vals, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))
            for fc in [2, 3, 4]:
                ws.cell(row=r_idx, column=fc).number_format = '#,##0.00 "ر.س"'

            net_cell = ws.cell(row=r_idx, column=4)
            if net < 0:
                net_cell.fill = _fill(CLR["negative_bg"])
                net_cell.font = _font(bold=True, color=CLR["negative"])
            else:
                net_cell.fill = _fill(CLR["positive_bg"])
                net_cell.font = _font(bold=True, color=CLR["positive"])

        last_data  = data_start + len(summary_df) - 1
        totals_row = last_data + 1
        ws.cell(row=totals_row, column=1, value="GRAND TOTAL")
        ws.cell(row=totals_row, column=2, value=f'=SUM(B{data_start}:B{last_data})')
        ws.cell(row=totals_row, column=3, value=f'=SUM(C{data_start}:C{last_data})')
        ws.cell(row=totals_row, column=4, value=f'=SUM(D{data_start}:D{last_data})')
        style_totals_row(ws, totals_row, num_cols)
        for fc in [2, 3, 4]:
            ws.cell(row=totals_row, column=fc).number_format = '#,##0.00 "ر.س"'

    set_column_widths(ws, {"A": 28, "B": 22, "C": 22, "D": 22})


def _sheet_revenue_expenses(wb: Workbook):
    """Sheet 2 — Revenue vs Expenses (monthly)."""
    ws = wb.create_sheet("Revenue vs Expenses")

    rev_exp_df = query_df(
        f"""
        SELECT
            TO_CHAR(transaction_date, 'YYYY-MM')              AS "Month",
            SUM(CASE WHEN txn_type='revenue'  THEN amount ELSE 0 END) AS "Revenue",
            SUM(CASE WHEN txn_type='expense'  THEN amount ELSE 0 END) AS "Expenses",
            SUM(CASE WHEN txn_type='revenue'  THEN amount ELSE -amount END) AS "Net Profit"
        FROM {TABLES['transactions']}
        GROUP BY TO_CHAR(transaction_date, 'YYYY-MM')
        ORDER BY 1 ASC
        """
    )

    num_cols = 4
    add_sheet_title(ws, "Revenue vs Expenses", f"Monthly Breakdown  |  Generated: {datetime.now():%Y-%m-%d %H:%M}", cols=num_cols)

    headers    = ["Month", "Revenue (ر.س)", "Expenses (ر.س)", "Net Profit (ر.س)"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, header_row, num_cols)

    data_start = header_row + 1

    if rev_exp_df.empty:
        ws.cell(row=data_start, column=1, value="No revenue/expense data found.")
        ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
    else:
        for r_idx, (_, row) in enumerate(rev_exp_df.iterrows(), start=data_start):
            net = float(row.get("Net Profit") or 0)
            vals = [
                row.get("Month", ""),
                float(row.get("Revenue")  or 0),
                float(row.get("Expenses") or 0),
                net,
            ]
            for c_idx, val in enumerate(vals, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))
            for fc in [2, 3, 4]:
                ws.cell(row=r_idx, column=fc).number_format = '#,##0.00 "ر.س"'

            net_cell = ws.cell(row=r_idx, column=4)
            if net < 0:
                net_cell.fill = _fill(CLR["negative_bg"])
                net_cell.font = _font(bold=True, color=CLR["negative"])
            else:
                net_cell.fill = _fill(CLR["positive_bg"])
                net_cell.font = _font(bold=True, color=CLR["positive"])

        last_data  = data_start + len(rev_exp_df) - 1
        totals_row = last_data + 1
        ws.cell(row=totals_row, column=1, value="TOTAL")
        for fc, col_letter in [(2, "B"), (3, "C"), (4, "D")]:
            ws.cell(row=totals_row, column=fc, value=f'=SUM({col_letter}{data_start}:{col_letter}{last_data})')
            ws.cell(row=totals_row, column=fc).number_format = '#,##0.00 "ر.س"'
        style_totals_row(ws, totals_row, num_cols)

        # ── Bar Chart ─────────────────────────────────────────────────────
        chart = BarChart()
        chart.type = "col"
        chart.title = "Revenue vs Expenses"
        chart.y_axis.title = "Amount (ر.س)"
        chart.x_axis.title = "Month"
        chart.grouping = "clustered"
        chart.width = 22
        chart.height = 14

        rev_ref = Reference(ws, min_col=2, min_row=header_row, max_row=last_data)
        exp_ref = Reference(ws, min_col=3, min_row=header_row, max_row=last_data)
        chart.add_data(rev_ref, titles_from_data=True)
        chart.add_data(exp_ref, titles_from_data=True)
        chart.series[0].graphicalProperties.solidFill = "2E86AB"
        chart.series[1].graphicalProperties.solidFill = "E63946"

        cats = Reference(ws, min_col=1, min_row=data_start, max_row=last_data)
        chart.set_categories(cats)
        ws.add_chart(chart, f"F{header_row}")

    set_column_widths(ws, {"A": 14, "B": 22, "C": 22, "D": 22})


def _sheet_clients_summary(wb: Workbook):
    """Sheet 3 — All Clients Summary."""
    ws = wb.create_sheet("Clients Summary")

    clients_df = query_df(
        f"""
        SELECT
            c.id                                              AS "Client ID",
            c.name                                            AS "Client Name",
            COUNT(t.id)                                       AS "Transactions",
            SUM(t.debit_amount)                               AS "Total Debits",
            SUM(t.credit_amount)                              AS "Total Credits",
            SUM(t.credit_amount - t.debit_amount)             AS "Net Balance"
        FROM {TABLES['clients']} c
        LEFT JOIN {TABLES['transactions']} t ON t.client_id = c.id
        GROUP BY c.id, c.name
        ORDER BY c.name
        """
    )

    num_cols = 6
    add_sheet_title(ws, "All Clients Summary", f"Generated: {datetime.now():%Y-%m-%d %H:%M}", cols=num_cols)

    headers    = ["Client ID", "Client Name", "Transactions", "Total Debits (ر.س)", "Total Credits (ر.س)", "Net Balance (ر.س)"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, header_row, num_cols)

    data_start = header_row + 1

    if clients_df.empty:
        ws.cell(row=data_start, column=1, value="No client data found.")
        ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
    else:
        for r_idx, (_, row) in enumerate(clients_df.iterrows(), start=data_start):
            net = float(row.get("Net Balance") or 0)
            vals = [
                row.get("Client ID"),
                row.get("Client Name", ""),
                int(row.get("Transactions") or 0),
                float(row.get("Total Debits")  or 0),
                float(row.get("Total Credits") or 0),
                net,
            ]
            for c_idx, val in enumerate(vals, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))
            for fc in [4, 5, 6]:
                ws.cell(row=r_idx, column=fc).number_format = '#,##0.00 "ر.س"'
            net_cell = ws.cell(row=r_idx, column=6)
            if net < 0:
                net_cell.fill = _fill(CLR["negative_bg"])
                net_cell.font = _font(bold=True, color=CLR["negative"])
            else:
                net_cell.fill = _fill(CLR["positive_bg"])
                net_cell.font = _font(bold=True, color=CLR["positive"])

        last_data  = data_start + len(clients_df) - 1
        totals_row = last_data + 1
        ws.cell(row=totals_row, column=1, value="TOTAL")
        ws.cell(row=totals_row, column=2, value=f"{len(clients_df)} Clients")
        ws.cell(row=totals_row, column=3, value=f'=SUM(C{data_start}:C{last_data})')
        ws.cell(row=totals_row, column=4, value=f'=SUM(D{data_start}:D{last_data})')
        ws.cell(row=totals_row, column=5, value=f'=SUM(E{data_start}:E{last_data})')
        ws.cell(row=totals_row, column=6, value=f'=SUM(F{data_start}:F{last_data})')
        style_totals_row(ws, totals_row, num_cols)
        for fc in [4, 5, 6]:
            ws.cell(row=totals_row, column=fc).number_format = '#,##0.00 "ر.س"'

    set_column_widths(ws, {"A": 12, "B": 30, "C": 16, "D": 22, "E": 22, "F": 22})


def _sheet_cashier_performance(wb: Workbook):
    """Sheet 4 — Cashier Performance."""
    ws = wb.create_sheet("Cashier Performance")

    cashier_df = query_df(
        f"""
        SELECT
            c.id                                              AS "Cashier ID",
            c.name                                            AS "Cashier Name",
            COUNT(DISTINCT t.shift)                           AS "Shifts Worked",
            SUM(CASE WHEN t.txn_type='collection' THEN t.amount ELSE 0 END) AS "Total Collected",
            SUM(CASE WHEN t.txn_type='payment'    THEN t.amount ELSE 0 END) AS "Total Paid Out",
            SUM(CASE WHEN t.txn_type='collection' THEN t.amount ELSE -t.amount END) AS "Net Cash"
        FROM {TABLES['cashiers']} c
        LEFT JOIN {TABLES['transactions']} t ON t.cashier_id = c.id
        GROUP BY c.id, c.name
        ORDER BY "Net Cash" DESC
        """
    )

    num_cols = 6
    add_sheet_title(ws, "Cashier Performance", f"Generated: {datetime.now():%Y-%m-%d %H:%M}", cols=num_cols)

    headers    = ["Cashier ID", "Cashier Name", "Shifts Worked", "Total Collected (ر.س)", "Total Paid Out (ر.س)", "Net Cash (ر.س)"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, header_row, num_cols)

    data_start = header_row + 1

    if cashier_df.empty:
        ws.cell(row=data_start, column=1, value="No cashier data found.")
        ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
    else:
        for r_idx, (_, row) in enumerate(cashier_df.iterrows(), start=data_start):
            net = float(row.get("Net Cash") or 0)
            vals = [
                row.get("Cashier ID"),
                row.get("Cashier Name", ""),
                int(row.get("Shifts Worked") or 0),
                float(row.get("Total Collected") or 0),
                float(row.get("Total Paid Out")  or 0),
                net,
            ]
            for c_idx, val in enumerate(vals, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))
            for fc in [4, 5, 6]:
                ws.cell(row=r_idx, column=fc).number_format = '#,##0.00 "ر.س"'
            net_cell = ws.cell(row=r_idx, column=6)
            if net < 0:
                net_cell.fill = _fill(CLR["negative_bg"])
                net_cell.font = _font(bold=True, color=CLR["negative"])
            else:
                net_cell.fill = _fill(CLR["positive_bg"])
                net_cell.font = _font(bold=True, color=CLR["positive"])

        last_data  = data_start + len(cashier_df) - 1
        totals_row = last_data + 1
        ws.cell(row=totals_row, column=1, value="TOTAL")
        ws.cell(row=totals_row, column=2, value=f"{len(cashier_df)} Cashiers")
        ws.cell(row=totals_row, column=3, value=f'=SUM(C{data_start}:C{last_data})')
        ws.cell(row=totals_row, column=4, value=f'=SUM(D{data_start}:D{last_data})')
        ws.cell(row=totals_row, column=5, value=f'=SUM(E{data_start}:E{last_data})')
        ws.cell(row=totals_row, column=6, value=f'=SUM(F{data_start}:F{last_data})')
        style_totals_row(ws, totals_row, num_cols)
        for fc in [4, 5, 6]:
            ws.cell(row=totals_row, column=fc).number_format = '#,##0.00 "ر.س"'

    set_column_widths(ws, {"A": 12, "B": 28, "C": 16, "D": 24, "E": 22, "F": 20})


def _sheet_supervisor_activity(wb: Workbook):
    """Sheet 5 — Supervisor Activity."""
    ws = wb.create_sheet("Supervisor Activity")

    sup_df = query_df(
        f"""
        SELECT
            s.id                                              AS "Supervisor ID",
            s.name                                            AS "Supervisor Name",
            COUNT(t.id)                                       AS "Total Actions",
            SUM(CASE WHEN t.status='approved' THEN 1 ELSE 0 END) AS "Approved",
            SUM(CASE WHEN t.status='rejected' THEN 1 ELSE 0 END) AS "Rejected",
            SUM(CASE WHEN t.status='pending'  THEN 1 ELSE 0 END) AS "Pending",
            SUM(t.amount)                                     AS "Total Amount Handled"
        FROM {TABLES['supervisors']} s
        LEFT JOIN {TABLES['transactions']} t ON t.supervisor_id = s.id
        GROUP BY s.id, s.name
        ORDER BY s.name
        """
    )

    num_cols = 7
    add_sheet_title(ws, "Supervisor Activity", f"Generated: {datetime.now():%Y-%m-%d %H:%M}", cols=num_cols)

    headers    = ["Supervisor ID", "Supervisor Name", "Total Actions", "Approved", "Rejected", "Pending", "Amount Handled (ر.س)"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    style_header_row(ws, header_row, num_cols)

    data_start = header_row + 1

    if sup_df.empty:
        ws.cell(row=data_start, column=1, value="No supervisor data found.")
        ws.merge_cells(f"A{data_start}:{get_column_letter(num_cols)}{data_start}")
    else:
        for r_idx, (_, row) in enumerate(sup_df.iterrows(), start=data_start):
            approved = int(row.get("Approved") or 0)
            rejected = int(row.get("Rejected") or 0)
            pending  = int(row.get("Pending")  or 0)
            vals = [
                row.get("Supervisor ID"),
                row.get("Supervisor Name", ""),
                int(row.get("Total Actions") or 0),
                approved,
                rejected,
                pending,
                float(row.get("Total Amount Handled") or 0),
            ]
            for c_idx, val in enumerate(vals, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            style_data_row(ws, r_idx, num_cols, alternate=(r_idx % 2 == 0))
            ws.cell(row=r_idx, column=7).number_format = '#,##0.00 "ر.س"'

            ws.cell(row=r_idx, column=4).fill = _fill(CLR["positive_bg"])
            ws.cell(row=r_idx, column=4).font = _font(bold=True, color=CLR["positive"])
            ws.cell(row=r_idx, column=5).fill = _fill(CLR["negative_bg"])
            ws.cell(row=r_idx, column=5).font = _font(bold=True, color=CLR["negative"])
            ws.cell(row=r_idx, column=6).fill = _fill("FFF3CD")
            ws.cell(row=r_idx, column=6).font = _font(bold=True, color="856404")

        last_data  = data_start + len(sup_df) - 1
        totals_row = last_data + 1
        ws.cell(row=totals_row, column=1, value="TOTAL")
        ws.cell(row=totals_row, column=2, value=f"{len(sup_df)} Supervisors")
        for fc, col in [(3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")]:
            ws.cell(row=totals_row, column=fc, value=f'=SUM({col}{data_start}:{col}{last_data})')
        style_totals_row(ws, totals_row, num_cols)
        ws.cell(row=totals_row, column=7).number_format = '#,##0.00 "ر.س"'

    set_column_widths(ws, {"A": 14, "B": 28, "C": 16, "D": 12, "E": 12, "F": 12, "G": 24})


def export_owner_report() -> int:
    print("\n── Exporting Owner Master Report ─────────────────────────────")

    wb = Workbook()
    # Remove the default blank sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    _sheet_balance_summary(wb)
    _sheet_revenue_expenses(wb)
    _sheet_clients_summary(wb)
    _sheet_cashier_performance(wb)
    _sheet_supervisor_activity(wb)

    out_path = OUTPUT_ROOT / "owner" / "owner_summary_report.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}  (5 sheets)")
    return 1


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                             ENTRY POINT                                 ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def main():
    print("=" * 62)
    print("  Raseed (رصيد) — Accounting Data Exporter")
    print(f"  Started: {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("=" * 62)

    # 1. Create output directory tree
    create_output_dirs()
    print(f"\n✓ Output directory ready: {OUTPUT_ROOT.resolve()}")

    # 2. Run all exporters
    counts = {}
    counts["clients"]     = export_clients()
    counts["cashiers"]    = export_cashiers()
    counts["supervisors"] = export_supervisors()
    counts["owner"]       = export_owner_report()

    # 3. Print summary table
    total = sum(counts.values())
    print("\n" + "=" * 62)
    print("  EXPORT SUMMARY")
    print("=" * 62)
    rows = [
        ("👤  Clients",     counts["clients"],     "reports/clients/"),
        ("🧾  Cashiers",    counts["cashiers"],    "reports/cashiers/"),
        ("👁   Supervisors", counts["supervisors"], "reports/supervisors/"),
        ("👑  Owner",       counts["owner"],       "reports/owner/"),
    ]
    for label, cnt, path in rows:
        bar = "█" * cnt if cnt <= 30 else ("█" * 30 + f"…(+{cnt-30})")
        print(f"  {label:<20} {cnt:>4} file(s)   {path}")
    print("-" * 62)
    print(f"  {'TOTAL':<20} {total:>4} file(s)")
    print("=" * 62)
    print(f"\n  All reports saved to: {(OUTPUT_ROOT).resolve()}\n")


if __name__ == "__main__":
    main()
